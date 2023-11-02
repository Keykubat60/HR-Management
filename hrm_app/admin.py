from django.contrib import admin
from .models import Unternehmen, Personal, Dokument, Abrechnung, Monatsabrechnung, Kind, Lohnprogramm
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import zipfile
from django.http import HttpResponse
from django.utils.html import format_html
import os
from datetime import datetime, timedelta
from django import forms
from django.forms import widgets
from django.db import models
from .forms import KindForm
from .forms import KindFormSet


class DokumentInline(admin.TabularInline):
    model = Dokument
    extra = 0  # Anzahl der leeren Formulare, die standardmäßig angezeigt werden


class AnmeldeStatusInline(admin.TabularInline):
    model = Lohnprogramm
    extra = 1  # Wie viele leere Formulare angezeigt werden sollen
    fields = ('status', 'datum')
    readonly_fields = ('datum',)


def export_xlsx(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields]

    # Manuell die Position von 'projekt' festlegen
    insert_index = field_names.index('standort') + 1
    field_names.insert(insert_index, 'projekt')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(meta)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Personal'

    # Add headers
    for col_num, column_title in enumerate(field_names, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = column_title

    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(field_names, 1):
            col_letter = get_column_letter(col_num)
            if field == 'standort':
                value = obj.standort.standort  # Zugriff auf das 'standort'-Feld des 'Unternehmen'-Objekts
            elif field == 'projekt':
                value = obj.standort.projekt if obj.standort else None  # Zugriff auf das 'projekt'-Feld des 'Unternehmen'-Objekts
            else:
                value = getattr(obj, field)
            ws['{}{}'.format(col_letter, row_num)] = value

    wb.save(response)

    return response


export_xlsx.short_description = "Ausgewählte Objekte als Excel exportieren"


class KindInline(admin.TabularInline):
    model = Kind
    formset = KindFormSet
    extra = 0


class PersonalAdmin(admin.ModelAdmin):
    inlines = [AnmeldeStatusInline, KindInline, DokumentInline]
    fieldsets = (
        ('Persönliche Informationen', {
            'fields': ('name', 'nachname', 'personalnummer', 'familienstand', 'staatsbuergerschaft', 'geburtsdatum', 'telefonnummer', 'email',
                       'email_passwort')
        }),
        ('Adresse', {
            'fields': ('strasse', 'postleitzahl', 'ort')
        }),
        ('Arbeitsdetails', {
            'fields': (
                'eintritt', 'austritt', 'status', ('vertragsart', 'sign'), 'standort')
        }),
        ('Finanzielle Informationen', {
            'fields': ('steuernummer', 'steuerklasse', 'krankenkassenname',
                       'kinderfreibetrag', 'sozialversicherungsnummer', 'iban', 'finanziell_komplett')
        }),
        ('Weitere Informationen', {
            'fields': (
                'transportmittel', 'uberaccount', 'ubereats_passwort',
                'gekuendigt', 'lp_abgemeldet', 'lp_angemeldet'
            )
        }),
    )
    list_filter = ('status', 'standort__projekt', 'standort__standort', 'finanziell_komplett',
                   'transportmittel')  # Aktualisierte Felder
    actions = [export_xlsx]
    list_display = (
        'name', 'nachname', 'personalnummer', 'status_colored', 'unternehmen_name', 'unternehmen_location',
        'transportmittel',
        'vertragsende', 'probezeit_status',
        'finanziell_komplett_colored', 'sign_colored')

    def status_colored(self, obj):
        color = 'green' if obj.status == 'aktiv' else 'red'
        return format_html('<span style="color: {};">{}</span>', color, obj.get_status_display())

    status_colored.admin_order_field = 'status'  # Erlaubt das Sortieren
    status_colored.short_description = 'Status'  # Setzt die Spaltenüberschrift

    def sign_colored(self, obj):
        if obj.sign:
            color = 'green'
            text = 'check'
        else:
            color = 'red'
            text = 'uncheck'
        return format_html('<span style="color: {};">{}</span>', color, text)

    sign_colored.admin_order_field = 'sign'  # Erlaubt das Sortieren nach diesem Feld
    sign_colored.short_description = 'Sign'  # Setzt die Spaltenüberschrift

    def unternehmen_name(self, obj):
        return obj.standort.projekt if obj.standort else 'Nicht zugewiesen'

    unternehmen_name.admin_order_field = 'standort__projekt'  # Erlaubt das Sortieren
    unternehmen_name.short_description = 'Projekt'

    def unternehmen_location(self, obj):
        return obj.standort.standort if obj.standort else 'Nicht zugewiesen'

    unternehmen_location.admin_order_field = 'standort__standort'  # Erlaubt das Sortieren
    unternehmen_location.short_description = 'Standort'

    def finanziell_komplett_colored(self, obj):
        if obj.finanziell_komplett:
            color = 'green'
            text = 'check'
        else:
            color = 'red'
            text = 'uncheck'
        return format_html('<span style="color: {};">{}</span>', color, text)

    finanziell_komplett_colored.admin_order_field = 'finanziell_komplett'  # Erlaubt das Sortieren
    finanziell_komplett_colored.short_description = 'Finanz'  # Setzt die Spaltenüberschrift

    def vertragsende(self, obj):
        if obj.austritt:
            delta = obj.austritt - datetime.now().date()
            days_remaining = delta.days
            if days_remaining <= 0:
                color = 'red'
                text = f'abgelaufen seit {days_remaining} Tage'
            elif days_remaining <= 30:
                color = 'orange'
                text = f'in {days_remaining} Tage'
            else:
                color = 'green'
                text = f'in {days_remaining} Tage'
            return format_html('<span style="color: {};">{}</span>', color, text)
        return 'Austrittsdatum nicht festgelegt'

    def probezeit_status(self, obj):
        if obj.eintritt:
            probezeit_ende = obj.eintritt + timedelta(days=180)  # 6 Monate = 180 Tage
            delta = probezeit_ende - datetime.now().date()
            days_remaining = delta.days
            color = ""
            if days_remaining <= 0:
                text = 'beendet'
            elif days_remaining <= 30:
                color = 'orange'
                text = f'noch {days_remaining} Tage'
            else:
                color = 'green'
                text = f'noch {days_remaining}'
            return format_html('<span style="color: {};">{}</span>', color, text)
        return 'Eintrittsdatum nicht festgelegt'

    vertragsende.admin_order_field = 'austritt'
    vertragsende.short_description = 'Vertragsende'

    probezeit_status.admin_order_field = 'eintritt'
    probezeit_status.short_description = 'Probezeit'


from django.urls import reverse
from django.utils.html import format_html


def download_as_zip(modeladmin, request, queryset):
    zip_filename = 'Dokumente.zip'
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={zip_filename}'

    with zipfile.ZipFile(response, 'w') as zipf:
        for doc in queryset:
            filename = os.path.basename(doc.datei.name)
            folder = f"{doc.personal.name}_{doc.personal.id}"
            zipf.write(doc.datei.path, os.path.join(folder, filename))

    return response


download_as_zip.short_description = "Ausgewählte Dokumente als ZIP herunterladen"


class DokumentAdmin(admin.ModelAdmin):
    list_display = ('file_name', 'personal', 'view_document')
    list_filter = ('personal__name',)
    actions = [download_as_zip]  # Fügen Sie die Aktion hier hinzu

    def file_name(self, obj):
        return os.path.basename(obj.datei.name)

    def view_document(self, obj):
        url = obj.datei.url
        return format_html(f'<a href="{url}" target="_blank">Öffnen</a>')

    view_document.short_description = 'Dokument anzeigen'
    file_name.short_description = 'Dateiname'


class PersonalChoiceField(forms.ModelChoiceField):
    def label_from_instance(self, obj):
        return f"{obj.name} {obj.nachname} {obj.personalnummer}"


class MonatsabrechnungAdmin(admin.ModelAdmin):
    list_display = ['monat', 'jahr']


from django.contrib.admin import SimpleListFilter


class FullNameListFilter(SimpleListFilter):
    title = 'Personal'
    parameter_name = 'personal'

    def lookups(self, request, model_admin):
        persons = set([c.personal for c in model_admin.model.objects.all()])
        return [(p.id, f"{p.name} {p.nachname}") for p in persons]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(personal__id__exact=self.value())
        else:
            return queryset


class AbrechnungAdmin(admin.ModelAdmin):
    list_display = ['monatsabrechnung', 'get_full_name', 'betrag', 'ueberwiesen', 'bar']
    change_list_template = 'admin/hrm_app/abrechnung/changelist_view.html'
    list_filter = ('monatsabrechnung__jahr', 'monatsabrechnung__monat', FullNameListFilter,)

    def get_full_name(self, obj):
        return f"{obj.personal.name} {obj.personal.nachname}"

    get_full_name.admin_order_field = 'personal__name'  # Erlaubt Sortierung
    get_full_name.short_description = 'Vollständiger Name'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'personal':
            monatsabrechnung_id = request.GET.get('monatsabrechnung')
            if monatsabrechnung_id:
                already_added = Abrechnung.objects.filter(monatsabrechnung_id=monatsabrechnung_id).values_list(
                    'personal', flat=True)
                kwargs['queryset'] = Personal.objects.exclude(id__in=already_added).filter(is_active=True)
            return PersonalChoiceField(queryset=kwargs.get('queryset', Personal.objects.all()))
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        form = super(AbrechnungAdmin, self).get_form(request, obj, **kwargs)
        form.base_fields['personal'].widget = widgets.Select(
            attrs={
                'onchange': '''
                    var iban = document.getElementById("id_iban");
                    var vertragsart = document.getElementById("id_vertragsart");
                    var selectedValue = this.options[this.selectedIndex].value;
                    if (selectedValue) {
                        fetch(`/api/get_personal_details/${selectedValue}/`)
                        .then(response => response.json())
                        .then(data => {
                            iban.innerHTML = "IBAN: " + data.iban;
                            vertragsart.innerHTML = "Vertragsart: " + data.vertragsart;
                        });
                    } else {
                        iban.innerHTML = "";
                        vertragsart.innerHTML = "";
                    }
                '''
            }
        )
        form.base_fields['monatsabrechnung'].widget = widgets.Select(
            attrs={
                'onchange': '''
                        var personsDiv = document.getElementById("id_persons_not_in_month");
                        var selectedValue = this.options[this.selectedIndex].value;
                        if (selectedValue) {
                            fetch(`/api/get_persons_not_in_month/${selectedValue}/`)
                            .then(response => response.json())
                            .then(data => {
                                var listHtml = "<ul>";
                                for (var i = 0; i < data.persons.length; i++) {
                                    var person = data.persons[i];
                                    listHtml += "<li>" + person.name + " " + person.nachname + " (" + person.personalnummer + ")</li>";
                                }
                                listHtml += "</ul>";
                                personsDiv.innerHTML = "Personen ohne Abrechnung für diesen Monat: " + listHtml;
                            });
                        } else {
                            personsDiv.innerHTML = "";
                        }
                    '''
            }
        )
        form.base_fields['betrag'].widget.attrs.update({
            'oninput': '''
                   var betrag = parseFloat(document.getElementById("id_betrag").value) || 0;
                   var ueberwiesen = parseFloat(document.getElementById("id_ueberwiesen").value) || 0;
                   var bar = parseFloat(document.getElementById("id_bar").value) || 0;
                   if (betrag === (ueberwiesen + bar)) {
                       document.getElementById("id_betrag").style.color = "green";
                   } else {
                       document.getElementById("id_betrag").style.color = "red";
                   }
               '''
        })
        form.base_fields['ueberwiesen'].widget.attrs.update({
            'oninput': 'document.getElementById("id_betrag").dispatchEvent(new Event("input"))'
        })
        form.base_fields['bar'].widget.attrs.update({
            'oninput': 'document.getElementById("id_betrag").dispatchEvent(new Event("input"))'
        })
        return form

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_additional_fields'] = True
        return super(AbrechnungAdmin, self).change_view(
            request, object_id, form_url, extra_context=extra_context,
        )

    def changelist_view(self, request, extra_context=None):
        # Grundlegende Ansicht
        response = super().changelist_view(request, extra_context)

        # Nur die Summe hinzufügen, wenn wir uns in der Ansicht 'changelist' befinden
        if hasattr(response, 'context_data') and 'cl' in response.context_data:
            queryset = response.context_data['cl'].queryset
            # Summe der Beträge
            total_betrag = queryset.aggregate(total=models.Sum('betrag'))['total'] or 0
            total_ueberwiesen = queryset.aggregate(total=models.Sum('ueberwiesen'))['total'] or 0
            total_bar = queryset.aggregate(total=models.Sum('bar'))['total'] or 0

            # Zusätzliche Kontextdaten hinzufügen
            if extra_context is None:
                extra_context = {}
            extra_context['total_betrag'] = total_betrag
            extra_context['total_ueberwiesen'] = total_ueberwiesen
            extra_context['total_bar'] = total_bar
            extra_context['show_color_coding'] = True  # Neue Kontextvariable

            response.context_data.update(extra_context)

        return response


@admin.register(Kind)
class KindAdmin(admin.ModelAdmin):
    list_display = ['name', 'geburtsdatum', 'personal']
    list_filter = ('personal__name',)
    form = KindForm

from django.utils.translation import gettext_lazy as _
from django.db.models import Subquery, OuterRef, Max
class AktuellerStatusFilter(admin.SimpleListFilter):
    title = _('aktueller Status')
    parameter_name = 'aktueller_status'

    def lookups(self, request, model_admin):
        return (
            ('angemeldet', _('Angemeldet')),
            ('abgemeldet', _('Abgemeldet')),
        )

    def queryset(self, request, queryset):
        # Subquery, um das neueste Datum und die neueste Uhrzeit für jedes Personal zu finden
        neuester_status_subquery = Lohnprogramm.objects.filter(
            personal=OuterRef('personal')
        ).values('personal').annotate(
            neuestes_datum=Max('datum')  # Stellen Sie sicher, dass 'datum' auch die Uhrzeit enthält
        ).values('id')  # Wir wählen die ID des neuesten Eintrags

        # Hauptquery, um den letzten Status basierend auf der ID des neuesten Eintrags zu filtern
        queryset = queryset.filter(
            id__in=neuester_status_subquery
        )

        if self.value() == 'angemeldet':
            return queryset.filter(status='angemeldet')
        if self.value() == 'abgemeldet':
            return queryset.filter(status='abgemeldet')
        return queryset
@admin.register(Lohnprogramm)
class LohnprogrammAdmin(admin.ModelAdmin):
    list_display = ['get_full_name', 'status_farbig', 'datum']
    list_filter = (FullNameListFilter,AktuellerStatusFilter)

    def get_full_name(self, obj):
        return f"{obj.personal.name} {obj.personal.nachname}"

    get_full_name.admin_order_field = 'personal__name'  # Erlaubt Sortierung
    get_full_name.short_description = 'Vollständiger Name'

    def status_farbig(self, obj):
        # Setzen Sie die Farbe basierend auf dem Status
        color = 'green' if obj.status == 'angemeldet' else 'red'
        return format_html(
            '<span style="color: {};">{}</span>',
            color,
            obj.get_status_display()
        )

    status_farbig.short_description = 'Status'


admin.site.register(Monatsabrechnung, MonatsabrechnungAdmin)

admin.site.register(Abrechnung, AbrechnungAdmin)
admin.site.register(Dokument, DokumentAdmin)
admin.site.register(Unternehmen)
admin.site.register(Personal, PersonalAdmin)
